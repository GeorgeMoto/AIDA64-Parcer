import os
import re
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from io import StringIO
import ttkbootstrap as ttk
import logging
import sys
from ui import AidaParserUI

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    filename='aida_parser.log',
    filemode='a'
)

# Константы
SECURITY_SOFTWARE = (
    "dallas",
    "dr.Web",
    "eset",
    "kaspersky",
    "nod",
    "secret net",
    "security Studio Endpoint Protection",
    "viPNet Client",
    "континент-АП",
    "криптоПро"
)


class HtmlParser:
    """Класс для парсинга HTML файлов"""

    def __init__(self, start_dir):
        self.start_dir = start_dir
        self.processed_files = set()  # Множество для отслеживания обработанных файлов

    def parse_html_file(self, file_path):
        """Парсинг HTML файла и извлечение нужной информации."""
        try:
            with open(file_path, 'r', encoding='Windows-1251', errors="ignore") as file:
                content = file.read()
        except UnicodeDecodeError:
            logging.error(f"Невозможно прочитать файл {file_path}. Проблемы с кодировкой.")
            return None

        soup = BeautifulSoup(content, 'html.parser')

        # Получаем относительный путь к файлу
        rel_path = os.path.relpath(file_path, start=self.start_dir)

        # Проверяем, не был ли файл уже обработан
        if rel_path in self.processed_files:
            logging.info(f"Файл {rel_path} уже был обработан. Пропускаем.")
            return None

        # Добавляем файл в множество обработанных
        self.processed_files.add(rel_path)

        # Анализ данных
        data = {
            'file_name': rel_path,
            'pc_name': self._extract_pc_name(soup),
            'os_info': self._extract_os_info(soup),
            'software': self._extract_software(soup),
            'users': self._extract_users(soup)
        }

        return data

    def _extract_pc_name(self, soup):
        """Извлечение имени компьютера"""
        rows = soup.find_all('tr')

        for row in rows:
            row_text = row.get_text(separator=' ', strip=True)

            if "Компьютер" in row_text:
                row_text = re.sub(r'\s+', ' ', row_text)
                os_match = re.search(r'Компьютер\s+(.*?)(?:\s+Генератор|\s*$)', row_text)

                if os_match:
                    os_string = os_match.group(1).strip()
                    logging.info(f"Найдена строка имя пк: {os_string}")
                    return os_string

        return "Имя компьютера не найдено"

    def _extract_os_info(self, soup):
        """Извлечение информации об операционной системе"""
        rows = soup.find_all('tr')


        for row in rows:
            row_text = row.get_text(separator=' ', strip=True)

            if "Операционная система" in row_text:
                row_text = re.sub(r'\s+', ' ', row_text)
                os_match = re.search(r'Операционная система\s+(.*?)(?:\s+Дата|\s*$)', row_text)

                if os_match:
                    os_string = os_match.group(1).strip()
                    logging.info(f"Найдена строка ОС: {os_string}")
                    return os_string

        return "Операционная система не найдена"

    def _extract_installed_programs(self, soup):
        """Функция для извлечения установленных программ."""
        program_table = soup.find("a", {"name": "installed programs"})  # Ищем якорь
        if program_table:
            program_table = program_table.find_next("table")  # Берем таблицу после заголовка
            table_html = str(program_table)  # Преобразуем таблицу в строку
        else:
            logging.warning("Таблица с установленными программами не найдена")
            return pd.DataFrame(columns=["Программа", "Версия"])

        try:
            df = pd.read_html(StringIO(table_html), header=0)[0]  # Берем первую таблицу
            df = df.iloc[1:, [2, 3]]  # Пропускаем первую строку, берем нужные столбцы
            df.columns = ["Программа", "Версия"]
            return df
        except Exception as e:
            logging.error(f"Ошибка при извлечении программ: {str(e)}")
            return pd.DataFrame(columns=["Программа", "Версия"])

    def _extract_software(self, soup):
        """Извлечение списка программного обеспечения"""
        software_list = []

        try:
            programs_df = self._extract_installed_programs(soup)
            for index, row in programs_df.iterrows():
                program_name = row["Программа"]
                program_version = row["Версия"]
                software_list.append(" ".join([str(program_name), str(program_version)]))
        except Exception as e:
            logging.error(f"Ошибка при обработке списка ПО: {str(e)}")

        return software_list

    def _extract_users(self, soup):
        """Извлечение списка пользователей"""
        users_list = []
        users_section = soup.find('a', attrs={'name': 'users'})
        users_count = 1

        if users_section:
            try:
                # Ищем все блоки dt, которые содержат имена пользователей
                user_blocks = users_section.parent.parent.find_next('table').find_all('td', class_='dt')
                for block in user_blocks:
                    user_text = block.text.strip()
                    # Извлекаем имя пользователя из строки [ UserName ]
                    user_match = re.search(r'\[\s*([^]]+)\s*\]', user_text)
                    if user_match:
                        user_name = user_match.group(1).strip()
                        users_list.append(f"{users_count}. {user_name} ")
                        users_count += 1
            except Exception as e:
                logging.error(f"Ошибка при извлечении пользователей: {str(e)}")

        return users_list


class ExcelReportGenerator:
    """Класс для создания и форматирования Excel отчетов"""

    def __init__(self):
        # Стиль границ
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        # Стиль заголовка
        self.header_font = Font(bold=True)
        self.header_alignment = Alignment(horizontal='center', vertical='center')
        # Стиль для текста с переносом
        self.wrap_alignment = Alignment(wrap_text=True, vertical='top')
        # Стиль для центрированного текста
        self.center_alignment = Alignment(vertical='top')

    def create_workbook(self, output_file):
        """Создает новую книгу Excel и настраивает заголовки"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Результаты анализа"

        # Установка альбомной ориентации
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

        # Добавляем заголовки
        headers = ["№", "Имя файла", "Тип ПК", "Операционная система", "Прикладное ПО", "Защитное ПО", "Пользователи"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header
            # Стиль заголовка
            ws[f"{col_letter}1"].font = self.header_font
            ws[f"{col_letter}1"].alignment = self.header_alignment

        # Настройка ширины столбцов
        ws.column_dimensions['A'].width = 5  # №
        ws.column_dimensions['B'].width = 40  # Имя файла
        ws.column_dimensions['C'].width = 20  # Тип ПК
        ws.column_dimensions['D'].width = 60  # Операционная система
        ws.column_dimensions['E'].width = 60  # Прикладное ПО
        ws.column_dimensions['F'].width = 60  # Защитное ПО
        ws.column_dimensions['G'].width = 40  # Пользователи

        return wb, ws

    def add_data_to_worksheet(self, ws, row_num, counter, data):
        """Добавляет данные в таблицу и форматирует ячейки"""
        ws[f"A{row_num}"] = counter
        ws[f"B{row_num}"] = data['file_name']
        ws[f"C{row_num}"] = data['pc_name']
        ws[f"D{row_num}"] = data['os_info']

        # Форматируем список ПО с нумерацией
        software_list = ""
        security_list = ""
        security_count = 1

        for i, sw in enumerate(data['software'], 1):
            software_list += f"{i}. {sw}\n"

            # Проверяем, является ли ПО защитным
            if self._contains_security_software(sw):
                security_list += f"{security_count}. {sw}\n"
                security_count += 1

        ws[f"E{row_num}"] = software_list
        ws[f"E{row_num}"].alignment = self.wrap_alignment

        ws[f"F{row_num}"] = security_list
        ws[f"F{row_num}"].alignment = self.wrap_alignment

        # Форматируем список пользователей
        users_list = "\n".join(data['users'])
        ws[f"G{row_num}"] = users_list

        # Применение стилей ячеек
        for col in range(1, 8):
            col_letter = get_column_letter(col)
            ws[f"{col_letter}{row_num}"].border = self.thin_border
            if col != 5 and col != 6:  # Не для колонок E и F, они уже имеют свой стиль
                ws[f"{col_letter}{row_num}"].alignment = self.center_alignment

        return row_num + 1

    def adjust_row_heights(self, ws, last_row):
        """Автоподбор высоты строк"""
        char_per_line = 60 // 7
        for row in range(2, last_row):
            cell_value = ws[f"E{row}"].value
            if cell_value:
                # Подсчитываем количество строк в содержимом ячейки
                line_count = sum(len(line) // char_per_line + 1 for line in cell_value.split("\n"))
                # Устанавливаем высоту строки пропорционально количеству строк
                ws.row_dimensions[row].height = max(30, min(line_count * 18, 1200))



    def _contains_security_software(self, software_name):
        """Проверяет, является ли программа защитным ПО."""
        lower_name = software_name.lower()
        for keyword in SECURITY_SOFTWARE:
            if keyword in lower_name:
                return True
        return False


def process_directory(directory, output_file, progress_callback=None):
    """Обрабатывает указанную директорию и создает Excel отчет.

    Args:
        directory (str): Путь к директории с HTML файлами
        output_file (str): Путь для сохранения Excel отчета
        progress_callback (function): Функция обратного вызова для обновления прогресса
    """
    # Инициализация классов
    parser = HtmlParser(directory)
    excel_generator = ExcelReportGenerator()

    # Создаем книгу Excel
    wb, ws = excel_generator.create_workbook(output_file)

    row_num = 2  # Начинаем с 2-й строки (после заголовков)
    counter = 1  # Счетчик для номера строки

    # Находим все .htm файлы в директории и поддиректориях
    html_files = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.lower().endswith('.htm') or file.lower().endswith('.html'):
                html_files.append(os.path.join(root, file))

    total_files = len(html_files)
    processed_files = 0

    # Сообщаем о начале обработки
    logging.info(f"Найдено {total_files} HTML файлов для обработки в {directory}")

    if progress_callback:
        progress_callback(0, total_files, "Начало обработки файлов...")

    # Обрабатываем каждый файл
    for file_path in html_files:
        try:
            logging.info(f"Обработка файла: {file_path}")

            # Обновляем прогресс
            if progress_callback:
                progress_callback(processed_files, total_files, f"Обработка: {os.path.basename(file_path)}")

            # Парсим HTML файл
            data = parser.parse_html_file(file_path)
            if data is None:
                processed_files += 1
                continue

            # Запись данных в Excel
            row_num = excel_generator.add_data_to_worksheet(ws, row_num, counter, data)
            counter += 1

        except Exception as e:
            logging.error(f"Ошибка при обработке файла {file_path}: {str(e)}", exc_info=True)

        processed_files += 1

        # Обновляем прогресс
        if progress_callback:
            progress_callback(processed_files, total_files)

    # Автоподбор высоты строк
    excel_generator.adjust_row_heights(ws, row_num)

    # Сохранение результатов
    wb.save(output_file)
    logging.info(f"Результаты сохранены в файл: {output_file}")

    # Финальное обновление прогресса
    if progress_callback:
        progress_callback(total_files, total_files, "Обработка завершена")

    return counter - 1  # Возвращаем количество обработанных файлов
